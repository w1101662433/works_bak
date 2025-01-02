import re, time, random, json, ast, logging, requests, asyncio, base64, platform, os, traceback, sys, datetime
from io import StringIO

# python3 -m pip install --upgrade pip
# python3 -m pip install requests lxml requests httpx flask sqlalchemy pymysql pyppeteer pyppeteer-stealth pandas numpy
# python3 -m pip install gevent gevent-websocket ws4py oss2 opencv-python==4.3.0.38


# from func_timeout import func_set_timeout
# func = func_set_timeout(interval)(func)
# -/\.(js|css|jpg|png|gif|svg|woff2|ico|woff)(\?|$)/
requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS = 'ALL'


class ddict(dict):
    def __init__(self, seq=None, **kwargs):
        # 多维字典：效率比setdefault略低，只有原生dict的1/4，海量数据时请不要使用
        if isinstance(seq, dict):
            for i in seq:
                new_seq = seq[i]
                if not isinstance(new_seq, dict): continue
                seq[i] = ddict(new_seq)
            dict.__init__(self, seq, **kwargs)
        elif seq:
            dict.__init__(self, seq, **kwargs)

    def __getitem__(self, item):
        if item in self:
            return dict.__getitem__(self, item)
        else:
            value = self[item] = type(self)()
            return value

    def __iadd__(self, other):
        if self:
            raise Exception('add forbidden, ddict is not None')
        return other


class SequentialGenerator:
    # 传入一个列表，可以无限迭代
    def __init__(self, li):
        self.li, self.loc = li, 0
        self.li_len = len(self.li)

    def __iter__(self):
        return self

    def __next__(self):
        resp = self.li[self.loc]
        self.loc += 1
        self.loc = 0 if self.loc == self.li_len else self.loc
        return resp


def unique_list(li):
    return list(dict.fromkeys(li))


def sort_dict(dic, by=0):
    return dict(sorted(dic.items(), key=lambda x: x[by]))


def cut_dict(dic, keys: list):
    return {k: dic.get(k) for k in keys}


def reverse_dict_key_values(dic):
    return {v: k for k, v in dic.items()}


def cut_text(text, length=159):
    t_li = re.findall('.{' + str(length) + '}', text)
    last_text = text[(len(t_li) * length):]
    last_text and t_li.append(last_text)
    return t_li


def split_list_to_n_part(old_li, n):
    length = len(old_li)
    cnt = length // n if length % n == 0 else length // n + 1
    return [old_li[i * cnt:(i + 1) * cnt] for i in range(0, n)]


def print_error_info(width=159, log=False):
    """用于打印报错信息"""
    fp = StringIO()
    traceback.print_exc(file=fp)
    message = fp.getvalue()
    message_list = sum([[i] if len(i) <= width else cut_text(i, width) for i in message.split('\n')], [])
    max_len = max([len(i) for i in message_list])
    message_list = map(lambda x: '|' + x.ljust(max_len) + '|', message_list)
    message = '+--TRY_EXCEPT:' + '-' * (max_len - 13) + '+\n' + '\n'.join(message_list) + '\n+' + '-' * max_len + '+'
    show_log(message) if log else print(message)
    return message


def get_random_str(length, special_chars=False):
    seed = r"1234567890abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if special_chars: seed += r"~!@#$%^&*()_+}{|:?><,./\;=-"
    return "".join([random.choice(seed) for i in range(length)])


def get_random_ip():
    return '.'.join([str(random.randint(0, 255)) for i in range(4)])


def get_host_ip():
    # 获取本机ip
    import socket
    ip = socket.gethostbyname(socket.gethostname())
    if ip.startswith('10.') or ip.startswith('168.') or ip.startswith('127.'):
        ip = retry_get(url='http://ifconfig.me').text
    return ip


def print_object_info(obj):
    try:
        print('__dict__.keys: ', obj.__dict__.keys())
        print('__dict__: ', obj.__dict__)
    except:
        print('obj has no __dict__')
    print('dir(): ', dir(obj))


def str_f_time(t=None, form="%Y-%m-%d %H:%M:%S"):
    t = t or time.time()
    t = str(t).strip()
    if not re.match(r'\d+\.?\d*$', t): return t
    t = float(t)
    if t > 10000000000: t = t // 1000
    return time.strftime(form, time.localtime(t))


def make_time(time_str):
    time_str = str(time_str).replace('T', ' ').strip()[:19]
    if time_str.isdigit() and len(time_str) > 9:
        return int(time_str[:9])
    element_dic = {19: "%Y-%m-%d %H:%M:%S", 16: "%Y-%m-%d %H:%M", 13: "%Y-%m-%d %H", 10: "%Y-%m-%d", 7: "%Y-%m", 8: "%Y%m%d"}
    form = element_dic.get(len(time_str))
    return int(time.mktime(time.strptime(time_str, form)))


def get_today_start(t=None):
    t = int(t) if t else int(time.time())
    return t - (t + 28800) % 86400


def get_delta_day(t=0):
    timedelta = datetime.timedelta
    now = datetime.datetime.fromtimestamp(t) if t else datetime.datetime.now()
    this_week_start = now - timedelta(days=now.weekday())
    this_month_start = datetime.datetime(now.year, now.month, 1)
    this_year_start = datetime.datetime(now.year, 1, 1)
    first_day_of_next_month = now.replace(day=1).replace(month=now.month + 1)
    this_month_end = first_day_of_next_month - datetime.timedelta(days=1)

    return {'this_week_start': get_today_start(this_week_start.timestamp()),
            'this_month_start': get_today_start(this_month_start.timestamp()),
            'this_year_start': get_today_start(this_year_start.timestamp()),
            'this_month_end': get_today_start(this_month_end.timestamp()), }


def seconds_to_time_long(seconds):
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    d, h = divmod(h, 24)
    return "%02d %02d:%02d:%02d" % (d, h, m, s)


def get_utc_time_str(t: int = None):
    t = t or int(time.time())
    res = datetime.datetime.utcfromtimestamp(t).isoformat("T")
    return res + '.000Z'


def utc_str_2_beijing_stamp(utc_str):
    # utc_str = '2023-01-12T06:37:12.300186Z'
    utc_str = utc_str.split('.')[0].replace('T', ' ')
    utc_stamp = make_time(utc_str)
    beijing_stamp = utc_stamp + 28800
    return beijing_stamp


def get_logger():
    logger = logging.getLogger(__file__)
    logger.setLevel(logging.DEBUG)
    ch = logging.StreamHandler()
    formatter = logging.Formatter("%(asctime)s %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    ch.setFormatter(formatter)
    logger.addHandler(ch)
    if platform.system() == 'Windows':
        project_name = os.path.abspath(__file__).split(os.sep)[-3]
        log_dir = r'C:/logs/' + str(project_name)
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        fh = logging.FileHandler(os.path.join(log_dir, str_f_time(form="%Y%m%d") + '.log'), encoding='utf-8')
        fh.setFormatter(formatter)
        logger.addHandler(fh)
    return logger


logger = get_logger()


def show_log(*args, **kwargs):
    logger.info(" ".join([str(i) for i in args]))


def show_error_log(*args, **kwargs):
    show_log("【ERROR】", *args)


def color_it(string, color='red'):
    color_trans = {'white': 30, 'red': 31, 'green': 32, 'yellow': 33, 'blue': 34, 'purple': 35, 'lightblue': 36, 'grey': 37}
    return "\033[0;{}m{}\033[0m".format(color_trans.get(color, 31), string)


def print_json_info(json_data, colorful=True):
    try:
        if isinstance(json_data, str):
            try:
                json_data = ast.literal_eval(json_data)
            except:
                json_data = json.loads(json_data)
        res = json.dumps(json_data, indent=2, ensure_ascii=False)
        if colorful:
            res = re.sub(r'"(.*?)":', '"' + color_it(r'\1') + '":', res)
            res = re.sub(r': "(.*?)"', ': "' + color_it(r'\1', 'blue') + '"', res)
            res = re.sub(r': (\d+)', ': ' + color_it(r'\1', 'green'), res)
            print(res)
    except:
        print('非json:', json_data)


def strip_html_tag(val):
    val = re.sub(r"<!--.*?-->", '', str(val))
    val = re.sub(r"&.{4};", ' ', val)
    return re.sub(r"<.*?>", '', val)


def url_replace_bracket(url):
    return url.replace('(', "%28").replace(')', "%29")


def match_jsonp(jsonp_str):
    return re.match(".*?({.*}).*", jsonp_str, re.S).group(1)


def retry_get(url, retry_times=3, session=None, interval=1, status_codes=None, error_keywords=None, **kwargs):
    return retry_request(url, 'GET', retry_times=retry_times, session=session, interval=interval, status_codes=status_codes,
                         error_keywords=error_keywords, **kwargs)


def retry_post(url, retry_times=3, session=None, interval=1, status_codes=None, **kwargs):
    return retry_request(url, 'POST', retry_times=retry_times, session=session, interval=interval, status_codes=status_codes, **kwargs)


def retry_request(url, method='GET', retry_times=3, session=None, interval=1, status_codes=None, error_keywords=None, **kwargs):
    for n_ in range(retry_times):
        try:
            if not kwargs.get('timeout'): kwargs['timeout'] = 30
            res = session.request(url=url, method=method, **kwargs) if session else requests.request(url=url, method=method, **kwargs)
            if status_codes and res.status_code not in status_codes:
                show_log('重试', 'res.status_code', res.status_code)
                time.sleep(interval)
                continue
            if error_keywords:
                for keyword in error_keywords:
                    if keyword in res.text:
                        raise Exception('found_error_keywords')
            return res
        except Exception as e:
            show_log('retry', str(e))
            show_log(url, method.lower(), n_)
            if n_ >= retry_times - 1:
                raise e

        time.sleep(interval)
    show_error_log('获取页面失败', url, 'kwargs', kwargs)
    raise Exception('获取页面失败')


def md5(string):
    import hashlib
    h = hashlib.md5()
    h.update(string.encode(encoding='utf-8'))
    return h.hexdigest()


def sha256(data, key):
    from hashlib import sha256
    import hmac
    key = key.encode('utf-8')
    message = data.encode('utf-8')
    return base64.b64encode(hmac.new(key, message, digestmod=sha256).digest()).decode()


def base_64(data):
    return base64.b64encode(str(data).encode()).decode()


def base_64_decode(data):
    return base64.b64decode(data).decode('utf-8')


def get_tmp_dir(root=False):
    if not root:
        pwd, _name = os.path.split(os.path.abspath(__file__))
        tmp_dir = os.path.join(pwd, '..', 'tmp')
    else:
        tmp_dir = '/tmp'
        if platform.system() == 'Windows':
            tmp_dir = r'c:\tmp'
    if not os.path.exists(tmp_dir): os.mkdir(tmp_dir)
    return tmp_dir


def get_cookies_list_from_session(session):
    return [{'name': i.name, 'value': i.value, 'domain': i.domain, 'path': i.path} for i in session.cookies]


def set_cookies_list_for_session(session, cookies_list):
    for cookie in cookies_list:
        tmp_dic = cut_dict(cookie, ['domain', 'name', 'value', 'path'])
        session.cookies.set(**tmp_dic)


def trans_dict_cookies_to_list(cookies: dict, domain):
    return [{'name': k, 'value': v, 'domain': domain, 'path': '/'} for k, v in cookies.items()]


def trans_list_cookies_to_dict(cookies: list):
    return {i['name']: i['value'] for i in cookies}


def dict_to_url_params(data: dict):
    from urllib.parse import quote
    return '&'.join([f"{k}={quote(v, safe='')}" for k, v in data.items()])


def get_tmp_json(file_name='cookies.json', default=None):
    tmp_file = os.path.join(get_tmp_dir(), file_name)
    none_resp = default or {}
    if not os.path.exists(tmp_file): return none_resp
    with open(tmp_file, 'r', encoding='utf-8') as f:
        data = f.read()
    if not data.strip(): return none_resp
    return json.loads(data)


def set_tmp_json(file_name='cookies.json', data: [dict, list] = None):
    tmp_file = os.path.join(get_tmp_dir(), file_name)
    tmp_dir = os.path.dirname(tmp_file)
    if not os.path.exists(tmp_dir):
        os.mkdir(tmp_dir)
    with open(tmp_file, 'w', encoding='utf-8') as f:
        f.write(json.dumps(data, indent=4, ensure_ascii=False))


def copy_msg_to_clipboard(text):
    import pyperclip
    pyperclip.copy(text)


def copy_image_to_clipboard(img_path: str):
    import win32clipboard
    from PIL import Image
    from io import BytesIO
    image = Image.open(img_path)
    output = BytesIO()
    image.save(output, 'BMP')
    data = output.getvalue()[14:]
    output.close()
    win32clipboard.OpenClipboard()
    win32clipboard.EmptyClipboard()
    win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
    win32clipboard.CloseClipboard()


def until_success(func, args=(), interval=1, retry_times=9, ):
    if not isinstance(args, tuple):
        raise Exception('args need be tuple')
    for _ in range(retry_times):
        try:
            return func(*args)
        except:
            time.sleep(interval)
            print_error_info(log=True)
    else:
        raise Exception('until_success failed')


async def until_success_async(func, args=(), interval=1, retry_times=9, ):
    if not isinstance(args, tuple):
        raise Exception('args need be tuple')
    for _ in range(retry_times):
        try:
            return await func(*args)
        except:
            await asyncio.sleep(interval)
            print_error_info(log=True)
    else:
        raise Exception('until_success failed')


def wave_num(num):
    wave = random.randint(10, 30)
    operator = random.choice(['+', '-'])
    num = (num + wave / 100) if operator == '+' else (num - wave / 100)
    return num


def clear_dead_python(key_words=None):
    key_words = key_words or ['python', 'chrome']
    if platform.system() == 'Windows': return
    for key_word in key_words:
        cmd = f'ps -ef | grep {key_word} | grep -v grep'
        res = os.popen(cmd)
        res = [i.split() for i in res.readlines()]
        for line in res:
            PPID, PID = line[2], line[1]
            if PPID == '1':
                os.popen(f'kill -9 {PID}')


def telnet_it(ip, port, timeout=0.5):
    import telnetlib
    server = telnetlib.Telnet()  # 创建一个Telnet对象
    try:
        server.open(ip, port, timeout=timeout)  # 利用Telnet对象的open方法进行tcp链接
        resp = True
    except:
        resp = False
    finally:
        server.close()
    return resp


class SSLFactory:
    def __init__(self):
        self.ORIGIN_CIPHERS = ('ECDH+AESGCM:DH+AESGCM:ECDH+AES256:DH+AES256:ECDH+AES128:DH+AES:ECDH+HIGH:'
                               'DH+HIGH:ECDH+3DES:DH+3DES:RSA+AESGCM:RSA+AES:RSA+HIGH:RSA+3DES')
        self.ciphers = self.ORIGIN_CIPHERS.split(":")

    def __call__(self):
        import ssl
        random.shuffle(self.ciphers)
        ciphers = ":".join(self.ciphers[random.randint(1, 4):])
        ciphers = ciphers + ":!aNULL:!eNULL:!MD5"
        context = ssl.create_default_context()
        context.set_ciphers(ciphers)
        return context


def get_random_ja3_str():
    def _shuffle(li, delete_count=0):
        random.shuffle(li)
        li = li[delete_count:]
        return '-'.join(li)

    ciphers_head = ['4865', '4866', '4867']
    ciphers_head = _shuffle(ciphers_head)
    ciphers_tail = '156-157-47-53'
    ciphers_body = ['49195-49199', '49196-49200', '52393-52392', '49171-49172']
    ciphers_body = _shuffle(ciphers_body, delete_count=1)
    ciphers = f"{ciphers_head}-{ciphers_body}-{ciphers_tail}"
    extensions = _shuffle('65281-18-27-43-0-5-51-13-11-17513-35-45-23-16-10-21'.split('-'))
    curves = '29-23-24'
    ja3_str = f"771,{ciphers},{extensions},{curves},0"
    show_log('获取ja3', ja3_str)
    return ja3_str


def get_pyhttpx_client(**kwargs):
    import pyhttpx
    session = pyhttpx.HttpSession(http2=True, ja3=get_random_ja3_str())
    return session


def get_cny_rate():
    url = 'https://www.ti.com/productmodel/currencyExchangeRates'
    res = retry_get(url=url)
    cny_rate = res.json().get('CNY')
    show_log('当前汇率为', cny_rate)
    return float(cny_rate)


def get_random_user_agent():
    if platform.system() == 'Windows':
        system = f'Windows NT {random.choice([7, 8, 10])}.{random.choice([0, 1])}; W{random.choice(["ow", "in"])}64; x64'
        centos = ''
    else:
        system = f'X1{random.choice([0, 1])}; Linux x86_64'
        centos1 = f'CentOS/{random.choice([6, 7, 8])}.{random.choice(range(1, 9))}'
        centos2 = f'Ubuntu/{random.choice([9, 10, 11])}.{random.choice(range(1, 9))}'
        centos = random.choice([centos1, centos2]) + ' '
    explorer = f'{random.choice(["Chromium", "Chrome"])}/{random.choice(range(86, 110))}.0.{random.choice(range(0, 4444))}.{random.choice(range(0, 186))}'
    pig_tail = f'537.{random.choice(range(19, 36))}'
    resp = f'Mozilla/5.0 ({system}) AppleWebKit/{pig_tail} (KHTML, like Gecko) {centos}{explorer} Safari/{pig_tail}'
    return resp


class RemoteConfig:
    def __init__(self, host='192.168.1.38', port=3306, username='root', password='Allone888', database='duo_ping_tai', table='dpt_config'):
        from sqlalchemy import create_engine
        from sqlalchemy.orm import sessionmaker
        engine = create_engine(f'mysql+pymysql://{username}:{password}@{host}:{port}/{database}?charset=utf8')
        self.table_name = table
        self.Session = sessionmaker(bind=engine)
        self.model = self.init_model_class()

    def init_model_class(self):
        from sqlalchemy import Column, Integer, String, Text
        from sqlalchemy.ext.declarative import declarative_base
        Base = declarative_base()

        def to_dict(_self):
            model_dict = dict(_self.__dict__)
            del model_dict['_sa_instance_state']
            return model_dict

        Base.to_dict = to_dict

        class DptConfig(Base):
            __tablename__ = self.table_name
            id = Column(Integer, primary_key=True, autoincrement=True)
            name = Column(String(128), unique=True)
            config = Column(Text())
            created = Column(Integer())

        return DptConfig

    def get_remote_config(self, name):
        session = self.Session()
        obj = session.query(self.model).filter(self.model.name == name).first()
        session.close()
        if not obj: return {}
        json_data = obj.to_dict()
        return json.loads(json_data.get('config', '{}'))


def split_xlsx_merged_cell(file_path, sheet_name, output_path='output.xlsx'):
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if isinstance(sheet_name, int):
        sheet = wb.worksheets[sheet_name]
    else:
        sheet_name_list = [i.title for i in wb.worksheets]
        sheet = wb.worksheets[sheet_name_list.index(sheet_name)]
    merged_cells_range = list(sheet.merged_cells)
    for merged_cell in merged_cells_range:
        min_row, min_col, max_row, max_col = merged_cell.min_row, merged_cell.min_col, merged_cell.max_row, merged_cell.max_col
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(merged_cell))
        for row in sheet.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.value = top_left_cell_value
    wb.save(output_path)


def read_multi_line_head_excel(file_path, header_line=(0, 1), sheet_name=None, **kwargs):
    import pandas as pd
    df = pd.read_excel(file_path, header=None, sheet_name=sheet_name, **kwargs)
    new_columns = df.iloc[header_line[0]].astype(str)
    if len(header_line) > 1:
        for line_count in header_line[1:]:
            new_columns = new_columns + '-' + df.iloc[line_count].astype(str)
    new_columns = [i.replace('\n', '').strip() for i in new_columns]
    df = df.iloc[header_line[-1] + 1:, ]
    df.columns = new_columns
    return df


def fill_df_date(df, date_col='time_str'):
    import pandas as pd
    df[date_col] = pd.to_datetime(df[date_col])
    end_date = pd.Timestamp(datetime.datetime.now().date())  # 使用当前日期
    date_range = pd.date_range(start=df[date_col].min(), end=end_date, freq='D')
    df.set_index(date_col, inplace=True)
    df = df.reindex(date_range, method='ffill')  # 使用前方数据填充
    df[date_col] = df.index
    return df


def get_column_trans_dict(model):
    from sqlalchemy import inspect
    return {column.name: column.comment for column in inspect(model).columns if column.name != 'id'}


def get_current_directory():
    absolute_file_path = os.path.abspath(__file__)
    return os.path.dirname(absolute_file_path)


def get_model_from_sqlalchemy_engine(engine, table_name="alarm_history"):
    from sqlalchemy.ext.automap import automap_base
    Base = automap_base()
    Base.prepare(engine)
    model = Base.classes.__getattr__(table_name)
    model.__tablename__ = table_name
    return model


def upsert(engine, table_name, df, unique_keys, model=None):
    from sqlalchemy.orm import sessionmaker
    db = get_model_from_sqlalchemy_engine(engine, table_name) if not model else model
    session = sessionmaker(bind=engine)()
    success_count = 0
    for _, item in df.iterrows():
        try:
            filter_by = [getattr(db, key) == item[key] for key in unique_keys]
            obj = session.query(db).filter(*filter_by).first() or db()
            for k, v in item.to_dict().items():
                obj.__setattr__(k, v)
            session.add(obj)
            session.commit()
            success_count += 1
        except Exception as e:
            show_error_log(e)
    session.close()
    return success_count, len(df)


def upsert_with_sql(engine, table_name, df, unique_keys):
    from sqlalchemy import text
    with engine.connect() as connection:
        for n_, item in df.iterrows():
            where_cond = ''
            for key in unique_keys:
                where_cond += f'and {key} = "{item[key]}" '
            where_cond = ' where ' + where_cond[4:]
            sql = f'select * from {table_name}' + where_cond
            data = connection.execute(text(sql)).fetchall()
            if not data:
                row_df = pd.DataFrame(item).T
                row_df.to_sql(con=engine, name=table_name, if_exists='append', index=False)


def table_to_df(engine, table_name):
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy import text
    session = sessionmaker(bind=engine)()
    result = session.execute(text(f"SELECT * FROM {table_name}"))
    df = pd.DataFrame(result.fetchall())
    df.columns = result.keys()
    session.close()
    return df


def print_return_value(func):
    def wrapper(*args, **kwargs):
        result = func(*args, **kwargs)
        show_log(f"【resp】{func.__name__!r}: {result!r}")
        return result

    return wrapper


if __name__ == "__main__":
    res = base_64_decode("aGVsbG8gd29ybGQh")
    print(res)
